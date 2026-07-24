from __future__ import annotations

import json
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook
import xlsxwriter


ROOT = Path(__file__).resolve().parent
DOWNLOADS = ROOT / "downloads"
OUTPUT = ROOT / "Business Lists Master - Deduped - 2026-07-05.xlsx"
STATS_OUTPUT = ROOT / "business_master_stats.json"


SOURCE_URLS = {
    "Home Improvement Retailers B2B.xlsx": "https://drive.google.com/file/d/15e-R2KzJkT_GTLCI2zyGhCodnjvGryhO",
    "Host scrapes.xlsx": "https://docs.google.com/spreadsheets/d/102PMJPvmbA8e-tQ7aasZ1B_4j-HQGcbzAe0s8WekYwo",
    "Mealscout Log.xlsx": "https://docs.google.com/spreadsheets/d/1-geeAc56PznDrXx5JdPZZKp4AToD8G0kCmNibMblyM4",
    "MealScout Round 1.xlsx": "https://docs.google.com/spreadsheets/d/1Jbb4h1YoZbor2gY1RxIDOB3_WZhqmSuC0ol7NVWpBnU",
    "Pensacola food trucks emails.xlsx": "https://docs.google.com/spreadsheets/d/1mdR3eofMlXatMgpvuwbKDdrxwHwbl1WTvQCmOjKNjMw",
    "Truck Owner Calls(5-30-26).xlsx": "https://docs.google.com/spreadsheets/d/100yNnK693VjXzEUQzheMcksjJNCgacDPDTHzV-MNgE4",
    "Bar Owner Calls.xlsx": "https://docs.google.com/spreadsheets/d/10Du7dmrwNyW8lKgPhPJm0TaCWVyC6Z1Z3PFnBt671l8",
    "DBPR_Active_Mobile_Vendors_Calling_Sheet.xlsx": "https://drive.google.com/file/d/1-lGs560XXJiDY3lpVaPF1cPTVoTspvJD",
    "DBPR_MFDV_Master_926.xlsx": "https://drive.google.com/file/d/1uh9xtFiEmArZxHaFBz-gvt0M8298VtfQ",
    "DBPR_MFDV_Escambia_174.xlsx": "https://drive.google.com/file/d/1F7-7dFNr8nKDuHyQUq1fszlnmbtz24CF",
    "DBPR_MFDV_SantaRosa_96.xlsx": "https://drive.google.com/file/d/18V2eU1IrDJzDtVtM51NQM2-MACpEigjj",
    "TradeScout Directory.xlsx": "https://docs.google.com/spreadsheets/d/10HegeRlL5Y8JLbOHjBF2E89b2pz6whAAAQF9KqeOn_w",
    "Property Management Leads.xlsx": "https://docs.google.com/spreadsheets/d/1OOXPtW5epTeErQbS5w9uVOiyL3IRr022Qh_2UnFa0Sw",
}

IDENTIFIED_NOT_IMPORTED = [
    {
        "source_file": "gulf_run",
        "source_url": "https://docs.google.com/spreadsheets/d/11AjN8tE9PgNGgIiXk5_vPtshZsUoFI71ahzRUiOzIvg",
        "status": "identified_not_imported",
        "notes": "Large raw scrape identified in Drive: about 298,887 rows x 11 columns. Left out of first import to keep the master under Google Sheets cell limits.",
    },
    {
        "source_file": "Escambia County Zip Code/Sunbiz",
        "source_url": "https://docs.google.com/spreadsheets/d/1nvtPBfJ1dWd9-ie7hY8f_ANQ3VB2fkrDkr3GoEM8-rU",
        "status": "identified_not_imported",
        "notes": "Sunbiz/business registry source identified in Drive. Not downloaded in this pass.",
    },
    {
        "source_file": "Territory Division MASTER SHEET",
        "source_url": "https://docs.google.com/spreadsheets/d/1cIMrENuKisGyb-249tOAugT4qYw8qePS_zWZPnQjk4M",
        "status": "identified_not_imported",
        "notes": "Operational territory/master sheet identified in Drive. Not downloaded in this pass.",
    },
    {
        "source_file": "Sales_Command_Center_CRM_with_31_Day_Content_Calendar",
        "source_url": "https://docs.google.com/spreadsheets/d/11xwbgHd9G_v0rd0322Q-TsfOyYliTT2fYv_CiltzANM",
        "status": "identified_not_imported",
        "notes": "CRM/operations workbook identified in Drive. Not treated as a raw scrape in this pass.",
    },
]

EXCLUDED_SCREENSHOT_SHEETS = [
    {
        "source_file": "MealScout Screenshot Processing Final Sheet 2026-06-09",
        "source_url": "https://docs.google.com/spreadsheets/d/1Qm7gwETnNlZNcXFJG5FuxhAWPyuKc7gqd02nyCBVf34",
        "reason": "Excluded per user: screenshot processing is separate from this.",
    },
    {
        "source_file": "My Copy of MealScout Screenshot Processing Final Sheet",
        "source_url": "https://docs.google.com/spreadsheets/d/1eSf1mC4tS4YaxagSo8bpXsvJ9Z2uowfQ5IXyMZjcYsk",
        "reason": "Excluded per user: screenshot processing is separate from this.",
    },
    {
        "source_file": "MealScout Screenshot Processing Corrected Rename Sheet 2026-06-09",
        "source_url": "https://docs.google.com/spreadsheets/d/1yOgl3h0S-ha63FlITmq6DI8JRmOgQcxTV8GlZZQHRUc",
        "reason": "Excluded per user: screenshot processing is separate from this.",
    },
    {
        "source_file": "MealScout Screenshot Processing Rename Sheet 2026-06-09",
        "source_url": "https://docs.google.com/spreadsheets/d/116iFlvmYGDKuqu044wo5QNfhQwFCtri4mmC7Ct4GO00",
        "reason": "Excluded per user: screenshot processing is separate from this.",
    },
]


HEADER_TERMS = {
    "business",
    "business_name",
    "businessname",
    "company",
    "company_name",
    "truck_name",
    "food_truck_name",
    "host_name",
    "name",
    "license",
    "licensenumber",
    "phone",
    "contact_phone",
    "email",
    "contact_email",
    "address",
    "fulladdress",
    "mailing_address",
    "city",
    "county",
    "state",
    "website",
    "instagram",
    "facebook",
    "category",
    "categories",
    "trade_type",
}

SKIP_DOMAINS = {
    "google.com",
    "www.google.com",
    "facebook.com",
    "www.facebook.com",
    "instagram.com",
    "www.instagram.com",
    "myfloridalicense.custhelp.com",
    "search.google.com",
}

GENERIC_NAME_REVIEW_SKIP = {
    "exxon",
    "shell",
    "chevron",
    "mobil",
    "circle k",
    "walmart",
    "wal mart",
    "publix",
    "aldi",
    "lowes",
    "lowe s",
    "home depot",
    "the home depot",
    "marathon",
    "bp",
}

CANONICAL_COLUMNS = [
    "canonical_id",
    "dedupe_status",
    "duplicate_count",
    "business_name",
    "normalized_name",
    "category",
    "source_family",
    "license_number",
    "status",
    "address",
    "city",
    "county",
    "state",
    "zip",
    "phone",
    "email",
    "website",
    "facebook",
    "instagram",
    "maps_url",
    "review_count",
    "rating",
    "source_files",
    "source_refs",
    "merge_group_id",
    "automatic_merge_basis",
    "notes",
]


def clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    text = text.replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    if text.lower() in {"nan", "none", "null"}:
        return ""
    return text


def ascii_fold(text: str) -> str:
    text = unicodedata.normalize("NFKD", text)
    return text.encode("ascii", "ignore").decode("ascii")


def normalize_header(value) -> str:
    text = ascii_fold(clean_text(value)).lower()
    text = text.replace("#", " number ")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def unique_headers(values) -> list[str]:
    headers = []
    seen = defaultdict(int)
    for idx, value in enumerate(values):
        name = normalize_header(value)
        if not name:
            name = f"column_{idx + 1}"
        seen[name] += 1
        if seen[name] > 1:
            name = f"{name}_{seen[name]}"
        headers.append(name)
    return headers


def header_score(values) -> int:
    score = 0
    for value in values:
        header = normalize_header(value)
        if header in HEADER_TERMS:
            score += 2
        elif any(term in header for term in HEADER_TERMS if len(term) > 4):
            score += 1
    return score


def manual_sheet_spec(file_name: str, sheet_name: str):
    if file_name == "MealScout Round 1.xlsx" and sheet_name == "Florida (Whole State)":
        return {
            "headers": [
                "business_name",
                "license_number",
                "phone",
                "contact_name",
                "address",
                "county",
                "state",
                "source_url",
                "column_9",
                "column_10",
                "column_11",
                "notes",
                "call_status",
            ],
            "start_row": 1,
            "notes": "No header row in source; applied MealScout vendor column map.",
        }
    if file_name == "Mealscout Log.xlsx" and sheet_name == "Florida (Whole State)":
        return {
            "headers": [
                "business_name",
                "license_number",
                "phone",
                "contact_name",
                "address",
                "county",
                "state",
                "source_url",
                "column_9",
                "column_10",
                "column_11",
                "notes",
                "call_status",
            ],
            "start_row": 1,
            "notes": "No header row in source; applied MealScout vendor column map.",
        }
    if file_name == "Mealscout Log.xlsx" and sheet_name == "Louisiana":
        return {
            "headers": [
                "id",
                "business_name",
                "address",
                "county",
                "state",
                "source",
                "source_link",
                "column_8",
                "column_9",
                "column_10",
                "column_11",
                "column_12",
                "column_13",
                "column_14",
                "column_15",
                "county_source",
                "state_source",
                "verified",
            ],
            "start_row": 1,
            "notes": "No header row in source; applied Louisiana state vendor column map.",
        }
    if file_name == "TradeScout Directory.xlsx" and sheet_name == "TradePartners":
        return {
            "headers": [
                "trade_type",
                "company_name",
                "phone",
                "website",
                "city",
                "state",
                "priority",
            ],
            "start_row": 1,
            "notes": "No header row in source; applied TradePartners column map.",
        }
    return None


def skip_sheet_reason(file_name: str, sheet_name: str) -> str:
    if file_name == "DBPR_Active_Mobile_Vendors_Calling_Sheet.xlsx" and sheet_name == "Calling Sheet":
        return "Skipped messy calling worksheet; clean Florida DBPR tab from same workbook was imported."
    if file_name == "Mealscout Log.xlsx" and sheet_name.strip() == "FB food truck groups":
        return "Skipped auxiliary Facebook group list; not a business lead table."
    if file_name == "Truck Owner Calls(5-30-26).xlsx":
        return "Skipped manual owner-call log because it does not contain business names or business contact fields."
    if sheet_name in {"Texas"}:
        return "Skipped empty sheet."
    return ""


def detect_headers(ws, file_name: str, sheet_name: str):
    manual = manual_sheet_spec(file_name, sheet_name)
    if manual:
        return manual["headers"], manual["start_row"], manual["notes"]

    rows = []
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        rows.append(row)

    best_row = None
    best_score = 0
    for idx, row in enumerate(rows, start=1):
        score = header_score(row)
        if score > best_score:
            best_score = score
            best_row = idx

    if best_row is None or best_score < 3:
        return [], 1, "No usable header row detected."

    headers = unique_headers(rows[best_row - 1])
    if file_name == "Mealscout Log.xlsx" and sheet_name == "Florida":
        for idx, header in enumerate(headers):
            if header.startswith("column_") and idx == 1:
                headers[idx] = "address"
    return headers, best_row + 1, f"Detected header row {best_row}."


def get_value(row: dict[str, str], *keys: str) -> str:
    for key in keys:
        key = normalize_header(key)
        if key in row and clean_text(row[key]):
            return clean_text(row[key])
    return ""


def first_nonempty(values) -> str:
    for value in values:
        text = clean_text(value)
        if text:
            return text
    return ""


def clean_phone(value: str) -> str:
    value = clean_text(value)
    if not value:
        return ""
    first = value.split(",")[0].strip()
    digits = re.sub(r"\D", "", first)
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) == 10:
        return f"({digits[0:3]}) {digits[3:6]}-{digits[6:10]}"
    if 7 <= len(digits) < 10:
        return digits
    return first


def phone_key(value: str) -> str:
    digits = re.sub(r"\D", "", clean_text(value))
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) >= 10:
        return digits[-10:]
    if len(digits) >= 7:
        return digits
    return ""


def clean_email(value: str) -> str:
    text = clean_text(value).lower()
    match = re.search(r"[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}", text)
    return match.group(0) if match else ""


def normalize_name(value: str) -> str:
    text = ascii_fold(clean_text(value)).lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    words = [word for word in text.split() if word]
    suffixes = {
        "llc",
        "inc",
        "incorporated",
        "corp",
        "corporation",
        "co",
        "company",
        "ltd",
        "limited",
        "pllc",
    }
    while words and words[-1] in suffixes:
        words.pop()
    return " ".join(words)


def normalize_address(value: str) -> str:
    text = ascii_fold(clean_text(value).replace("\u2014", " "))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_city_state_zip(value: str):
    text = clean_text(value)
    if not text:
        return "", "", ""
    patterns = [
        r"(?P<city>[A-Za-z .'\-]+),\s*(?P<state>[A-Z]{2})\s*(?P<zip>\d{5}(?:-\d{4})?)?$",
        r"\b(?P<city>[A-Za-z .'\-]+)\s+(?P<state>[A-Z]{2})\s+(?P<zip>\d{5}(?:-\d{4})?)$",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            city = clean_text(match.groupdict().get("city", ""))
            state = clean_text(match.groupdict().get("state", ""))
            zip_code = clean_text(match.groupdict().get("zip", ""))
            return city, state, zip_code
    return "", "", ""


def domain_from_url(value: str) -> str:
    text = clean_text(value).strip()
    if not text or "search.google.com" in text or "google.com/search" in text:
        return ""
    if "@" in text and not text.startswith("http"):
        return ""
    candidate = text if re.match(r"^[a-z]+://", text, flags=re.I) else f"https://{text}"
    try:
        parsed = urlparse(candidate)
    except ValueError:
        return ""
    domain = parsed.netloc.lower().split("@")[-1].split(":")[0]
    domain = domain[4:] if domain.startswith("www.") else domain
    if not domain or "." not in domain or domain in SKIP_DOMAINS:
        return ""
    return domain


def maps_cid(value: str) -> str:
    text = clean_text(value)
    if not text:
        return ""
    parsed = urlparse(text)
    query = parse_qs(parsed.query)
    cid = first_nonempty(query.get("cid", []))
    if cid:
        return cid
    match = re.search(r"cid=(\d+)", text)
    return match.group(1) if match else ""


def social_key(value: str, platform: str) -> str:
    text = clean_text(value)
    if not text:
        return ""
    lowered = text.lower()
    if "search" in lowered or "explore/search" in lowered:
        return ""
    if text.startswith("@"):
        handle = text[1:]
    elif platform in lowered and "://" in lowered:
        parsed = urlparse(text)
        parts = [part for part in parsed.path.split("/") if part]
        if not parts:
            return ""
        handle = parts[0]
    else:
        handle = text
    handle = re.sub(r"[^a-zA-Z0-9._-]+", "", handle).strip("._-").lower()
    if handle in {"pages", "groups", "share", "search", "explore"}:
        return ""
    return handle


def source_family(file_name: str, sheet_name: str, row: dict[str, str]) -> str:
    if file_name == "Home Improvement Retailers B2B.xlsx":
        if sheet_name == "Sheet1":
            return "contractor_license"
        if sheet_name == "Sheet2":
            return "contractor_osm"
        if sheet_name == "Sheet3":
            return "contractor_maps"
    if file_name == "TradeScout Directory.xlsx":
        return "tradescout_contractors"
    if file_name.startswith("DBPR") or file_name.startswith("MealScout") or file_name == "Pensacola food trucks emails.xlsx":
        return "food_vendor"
    if file_name == "Mealscout Log.xlsx":
        if sheet_name in {"Hosts", "Possible Hosts", "Florida"}:
            return "host_lead"
        return "food_vendor"
    if file_name == "Bar Owner Calls.xlsx":
        return "bar_host"
    if file_name == "Host scrapes.xlsx":
        return "host_lead"
    if file_name == "Property Management Leads.xlsx":
        return "property_management"
    return "business_lead"


def usable_name(name: str) -> bool:
    text = clean_text(name)
    if len(text) < 2:
        return False
    header_like = {"name", "business name", "company name", "truck name", "host name"}
    if text.lower() in header_like:
        return False
    if re.fullmatch(r"[\W_]+", text):
        return False
    return True


def row_to_record(
    row: dict[str, str],
    file_name: str,
    sheet_name: str,
    row_number: int,
    record_id: str,
):
    name = get_value(
        row,
        "business_name",
        "businessname",
        "business",
        "company_name",
        "company",
        "truck_name",
        "food_truck_name",
        "host_name",
        "name",
    )
    if not usable_name(name):
        return None

    license_number = get_value(row, "license_number", "licensenumber", "license")
    category = get_value(
        row,
        "categories",
        "category",
        "category_label",
        "categorylabel",
        "classification",
        "license_type",
        "cuisine_type",
        "property_type",
        "trade_type",
    )
    status = get_value(row, "status", "operating_status", "stage", "disposition", "approved", "verified", "call_status")
    address = get_value(
        row,
        "fulladdress",
        "full_address",
        "mailing_address",
        "address",
        "location",
        "street_license_location",
        "street",
        "column_2",
    )
    municipality = get_value(row, "municipality")
    if address and municipality and municipality.lower() not in address.lower():
        address = f"{address}, {municipality}"

    city = get_value(row, "city", "municipality", "service_area_city_county", "service_area", "area")
    state = get_value(row, "state", "sourcestate", "source_state", "state_source")
    zip_code = get_value(row, "zip", "postcode")
    county = get_value(row, "county", "county_source")

    for candidate in [city, address, municipality]:
        parsed_city, parsed_state, parsed_zip = parse_city_state_zip(candidate)
        if parsed_city and (not city or "," in city or re.search(r"\d{5}", city)):
            city = parsed_city
        if parsed_state and not state:
            state = parsed_state
        if parsed_zip and not zip_code:
            zip_code = parsed_zip

    phone = clean_phone(get_value(row, "phone", "contact_phone", "number", "phones"))
    email = clean_email(get_value(row, "email", "contact_email"))
    website = get_value(row, "website")
    facebook = get_value(row, "facebook", "facebook_page")
    instagram = get_value(row, "instagram")
    maps_url = get_value(row, "google_maps_url", "maps_url", "map_url")
    review_count = get_value(row, "review_count", "reviews")
    rating = get_value(row, "average_rating", "rating")
    notes = get_value(
        row,
        "notes",
        "source_notes",
        "report",
        "calls",
        "response_status",
        "objection",
        "priority",
    )

    family = source_family(file_name, sheet_name, row)
    if not category:
        if family.startswith("contractor") or family == "tradescout_contractors":
            category = "contractor / home improvement"
        elif family == "food_vendor":
            category = "food vendor / restaurant"
        elif family in {"host_lead", "bar_host"}:
            category = "host lead"
        elif family == "property_management":
            category = "property management"

    return {
        "record_id": record_id,
        "business_name": name,
        "normalized_name": normalize_name(name),
        "category": category,
        "source_family": family,
        "source_file": file_name,
        "source_url": SOURCE_URLS.get(file_name, ""),
        "source_tab": sheet_name,
        "source_row": row_number,
        "license_number": license_number,
        "status": status,
        "address": address,
        "normalized_address": normalize_address(address),
        "city": clean_text(city).title() if clean_text(city).isupper() else clean_text(city),
        "county": county,
        "state": clean_text(state).upper() if len(clean_text(state)) == 2 else clean_text(state),
        "zip": zip_code,
        "phone": phone,
        "phone_key": phone_key(phone),
        "email": email,
        "website": website,
        "website_domain": domain_from_url(website),
        "facebook": facebook,
        "facebook_key": social_key(facebook, "facebook"),
        "instagram": instagram,
        "instagram_key": social_key(instagram, "instagram"),
        "maps_url": maps_url,
        "maps_cid": maps_cid(maps_url),
        "review_count": review_count,
        "rating": rating,
        "notes": notes[:500],
    }


def strong_keys(record: dict[str, str]) -> list[str]:
    keys = []
    license_clean = re.sub(r"[^A-Za-z0-9]", "", record.get("license_number", "")).upper()
    if len(license_clean) >= 4:
        keys.append(f"license:{license_clean}")
    if record.get("phone_key"):
        keys.append(f"phone:{record['phone_key']}")
    if record.get("email"):
        keys.append(f"email:{record['email']}")
    if record.get("maps_cid"):
        keys.append(f"maps:{record['maps_cid']}")
    if record.get("facebook_key"):
        keys.append(f"facebook:{record['facebook_key']}")
    if record.get("instagram_key"):
        keys.append(f"instagram:{record['instagram_key']}")
    if record.get("normalized_name") and record.get("normalized_address"):
        if len(record["normalized_name"]) >= 5 and len(record["normalized_address"]) >= 8:
            keys.append(f"name_address:{record['normalized_name']}|{record['normalized_address']}")
    return keys


class UnionFind:
    def __init__(self, size: int):
        self.parent = list(range(size))
        self.rank = [0] * size

    def find(self, item: int) -> int:
        while self.parent[item] != item:
            self.parent[item] = self.parent[self.parent[item]]
            item = self.parent[item]
        return item

    def union(self, left: int, right: int):
        root_left = self.find(left)
        root_right = self.find(right)
        if root_left == root_right:
            return
        if self.rank[root_left] < self.rank[root_right]:
            root_left, root_right = root_right, root_left
        self.parent[root_right] = root_left
        if self.rank[root_left] == self.rank[root_right]:
            self.rank[root_left] += 1


def quality_score(record: dict[str, str]) -> int:
    score = 0
    for field in [
        "license_number",
        "phone",
        "email",
        "website",
        "facebook",
        "instagram",
        "address",
        "city",
        "county",
        "state",
        "category",
    ]:
        if record.get(field):
            score += 1
    if record.get("source_family") in {"contractor_license", "food_vendor"}:
        score += 2
    if record.get("source_family") in {"contractor_maps", "host_lead"}:
        score += 1
    return score


def combine_unique(values, limit: int | None = None) -> str:
    seen = []
    for value in values:
        text = clean_text(value)
        if text and text not in seen:
            seen.append(text)
    if limit is not None and len(seen) > limit:
        kept = seen[:limit]
        kept.append(f"+{len(seen) - limit} more")
        return "; ".join(kept)
    return "; ".join(seen)


def build_records():
    records = []
    sheet_stats = []
    record_number = 0

    for path in sorted(DOWNLOADS.glob("*.xlsx")):
        file_name = path.name
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            sheet_stats.append(
                {
                    "source_file": file_name,
                    "source_url": SOURCE_URLS.get(file_name, ""),
                    "sheet": "",
                    "status": "error",
                    "rows_scanned": 0,
                    "records_imported": 0,
                    "notes": f"{type(exc).__name__}: {exc}",
                }
            )
            continue

        for ws in wb.worksheets:
            sheet_name = ws.title
            reason = skip_sheet_reason(file_name, sheet_name)
            if reason:
                sheet_stats.append(
                    {
                        "source_file": file_name,
                        "source_url": SOURCE_URLS.get(file_name, ""),
                        "sheet": sheet_name,
                        "status": "skipped",
                        "rows_scanned": 0,
                        "records_imported": 0,
                        "notes": reason,
                    }
                )
                continue

            headers, start_row, note = detect_headers(ws, file_name, sheet_name)
            if not headers:
                sheet_stats.append(
                    {
                        "source_file": file_name,
                        "source_url": SOURCE_URLS.get(file_name, ""),
                        "sheet": sheet_name,
                        "status": "skipped",
                        "rows_scanned": 0,
                        "records_imported": 0,
                        "notes": note,
                    }
                )
                continue

            scanned = 0
            imported = 0
            max_col = len(headers)
            for row_number, values in enumerate(
                ws.iter_rows(min_row=start_row, max_col=max_col, values_only=True),
                start=start_row,
            ):
                scanned += 1
                row = {
                    headers[idx]: clean_text(values[idx]) if idx < len(values) else ""
                    for idx in range(len(headers))
                }
                if not any(row.values()):
                    continue
                record_number += 1
                record = row_to_record(row, file_name, sheet_name, row_number, f"R{record_number:07d}")
                if record:
                    records.append(record)
                    imported += 1

            sheet_stats.append(
                {
                    "source_file": file_name,
                    "source_url": SOURCE_URLS.get(file_name, ""),
                    "sheet": sheet_name,
                    "status": "imported" if imported else "no_business_rows",
                    "rows_scanned": scanned,
                    "records_imported": imported,
                    "notes": note,
                }
            )
        wb.close()

    for item in IDENTIFIED_NOT_IMPORTED:
        sheet_stats.append(
            {
                "source_file": item["source_file"],
                "source_url": item["source_url"],
                "sheet": "",
                "status": item["status"],
                "rows_scanned": "",
                "records_imported": "",
                "notes": item["notes"],
            }
        )

    return records, sheet_stats


def dedupe_records(records: list[dict[str, str]]):
    union_find = UnionFind(len(records))
    key_owner = {}
    record_keys = []

    for idx, record in enumerate(records):
        keys = strong_keys(record)
        record_keys.append(keys)
        for key in keys:
            if key in key_owner:
                union_find.union(idx, key_owner[key])
            else:
                key_owner[key] = idx

    groups = defaultdict(list)
    for idx in range(len(records)):
        groups[union_find.find(idx)].append(idx)

    canonical_rows = []
    duplicate_groups = []
    duplicate_members = []
    idx_to_canonical = {}
    sorted_groups = sorted(
        groups.values(),
        key=lambda indexes: (
            normalize_name(records[indexes[0]].get("business_name", "")),
            records[indexes[0]].get("city", ""),
            indexes[0],
        ),
    )

    merge_counter = 0
    for canonical_number, indexes in enumerate(sorted_groups, start=1):
        group_records = [records[idx] for idx in indexes]
        group_records.sort(key=lambda record: (-quality_score(record), record["source_file"], record["source_tab"]))
        best = group_records[0]
        group_keys = []
        for idx in indexes:
            group_keys.extend(record_keys[idx])
        basis = combine_unique(group_keys, limit=12)
        merge_group_id = ""
        if len(indexes) > 1:
            merge_counter += 1
            merge_group_id = f"MG{merge_counter:06d}"

        canonical_id = f"BUS{canonical_number:07d}"
        for idx in indexes:
            idx_to_canonical[idx] = canonical_id

        def choose(field: str) -> str:
            return first_nonempty([record.get(field, "") for record in group_records])

        source_refs = [
            f"{record['source_file']}::{record['source_tab']}#{record['source_row']}"
            for record in group_records
        ]
        row = {
            "canonical_id": canonical_id,
            "dedupe_status": "merged" if len(indexes) > 1 else "unique",
            "duplicate_count": len(indexes),
            "business_name": choose("business_name"),
            "normalized_name": choose("normalized_name"),
            "category": combine_unique([record.get("category", "") for record in group_records], limit=5),
            "source_family": combine_unique([record.get("source_family", "") for record in group_records], limit=5),
            "license_number": combine_unique([record.get("license_number", "") for record in group_records], limit=5),
            "status": combine_unique([record.get("status", "") for record in group_records], limit=5),
            "address": choose("address"),
            "city": choose("city"),
            "county": choose("county"),
            "state": choose("state"),
            "zip": choose("zip"),
            "phone": choose("phone"),
            "email": choose("email"),
            "website": choose("website"),
            "facebook": choose("facebook"),
            "instagram": choose("instagram"),
            "maps_url": choose("maps_url"),
            "review_count": choose("review_count"),
            "rating": choose("rating"),
            "source_files": combine_unique([record.get("source_file", "") for record in group_records], limit=8),
            "source_refs": combine_unique(source_refs, limit=8),
            "merge_group_id": merge_group_id,
            "automatic_merge_basis": basis,
            "notes": combine_unique([record.get("notes", "") for record in group_records], limit=3)[:500],
        }
        canonical_rows.append(row)

        if len(indexes) > 1:
            duplicate_groups.append(
                {
                    "merge_group_id": merge_group_id,
                    "group_size": len(indexes),
                    "primary_business_name": row["business_name"],
                    "city": row["city"],
                    "state": row["state"],
                    "automatic_merge_basis": basis,
                    "record_ids": combine_unique([record.get("record_id", "") for record in group_records], limit=25),
                    "source_files": row["source_files"],
                    "phones": combine_unique([record.get("phone", "") for record in group_records], limit=8),
                    "emails": combine_unique([record.get("email", "") for record in group_records], limit=8),
                    "licenses": combine_unique([record.get("license_number", "") for record in group_records], limit=8),
                    "names_seen": combine_unique([record.get("business_name", "") for record in group_records], limit=8),
                }
            )
            for record in group_records:
                duplicate_members.append(
                    {
                        "merge_group_id": merge_group_id,
                        "record_id": record["record_id"],
                        "business_name": record["business_name"],
                        "source_file": record["source_file"],
                        "source_tab": record["source_tab"],
                        "source_row": record["source_row"],
                        "license_number": record.get("license_number", ""),
                        "phone": record.get("phone", ""),
                        "email": record.get("email", ""),
                        "address": record.get("address", ""),
                        "city": record.get("city", ""),
                        "state": record.get("state", ""),
                    }
                )

    review_rows = []
    medium_groups = defaultdict(list)
    for idx, record in enumerate(records):
        name = record.get("normalized_name", "")
        city = clean_text(record.get("city", "")).lower()
        state = clean_text(record.get("state", "")).upper()
        if len(name) < 4 or not city or name in GENERIC_NAME_REVIEW_SKIP:
            continue
        medium_groups[f"{name}|{city}|{state}"].append(idx)

    review_group_counter = 0
    for key, indexes in sorted(medium_groups.items()):
        canonical_ids = {idx_to_canonical[idx] for idx in indexes}
        if len(indexes) < 2 or len(canonical_ids) < 2:
            continue
        review_group_counter += 1
        review_group_id = f"RV{review_group_counter:06d}"
        for idx in indexes[:20]:
            record = records[idx]
            review_rows.append(
                {
                    "review_group_id": review_group_id,
                    "reason": "Same normalized name and city/state, but no automatic strong key match.",
                    "canonical_id": idx_to_canonical[idx],
                    "record_id": record["record_id"],
                    "business_name": record["business_name"],
                    "city": record.get("city", ""),
                    "state": record.get("state", ""),
                    "phone": record.get("phone", ""),
                    "email": record.get("email", ""),
                    "address": record.get("address", ""),
                    "source_file": record["source_file"],
                    "source_tab": record["source_tab"],
                    "source_row": record["source_row"],
                }
            )
        if len(review_rows) >= 10000:
            break

    return canonical_rows, duplicate_groups, duplicate_members, review_rows


def write_sheet(workbook, name: str, columns: list[str], rows: list[dict[str, str]], widths: dict[str, int] | None = None):
    worksheet = workbook.add_worksheet(name[:31])
    header_fmt = workbook.add_format({"bold": True, "bg_color": "#D9EAF7", "border": 1})
    for col_idx, column in enumerate(columns):
        worksheet.write(0, col_idx, column, header_fmt)
        width = widths.get(column, 18) if widths else 18
        worksheet.set_column(col_idx, col_idx, width)
    worksheet.freeze_panes(1, 0)
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, column in enumerate(columns):
            worksheet.write(row_idx, col_idx, row.get(column, ""))
    if rows:
        worksheet.autofilter(0, 0, len(rows), len(columns) - 1)


def write_workbook(
    canonical_rows,
    duplicate_groups,
    duplicate_members,
    review_rows,
    sheet_stats,
    raw_count: int,
):
    workbook = xlsxwriter.Workbook(
        OUTPUT,
        {"constant_memory": True, "strings_to_urls": False, "nan_inf_to_errors": True},
    )
    title_fmt = workbook.add_format({"bold": True, "font_size": 14})
    key_fmt = workbook.add_format({"bold": True, "bg_color": "#E2F0D9"})

    readme = workbook.add_worksheet("README")
    readme.set_column(0, 0, 28)
    readme.set_column(1, 1, 110)
    readme.write(0, 0, "Business Lists Master", title_fmt)
    readme_rows = [
        ("Created", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Original files", "Left untouched. This workbook is a new organized/deduped output."),
        ("Screenshot processing", "Excluded per user request; see Excluded Screenshot Sheets tab."),
        ("Raw business rows imported", raw_count),
        ("Canonical business rows", len(canonical_rows)),
        ("Automatic duplicate groups", len(duplicate_groups)),
        ("Duplicate member rows", len(duplicate_members)),
        ("Needs-review rows", len(review_rows)),
        (
            "Automatic merge keys",
            "License number, phone, email, Google Maps CID, direct Facebook/Instagram handle, and exact normalized name + address.",
        ),
        (
            "Needs-review rule",
            "Same normalized name and city/state without a strong key match. These were not automatically merged.",
        ),
        (
            "Cell-limit note",
            "Very large Drive sources identified but not downloaded are listed in Source Inventory so they can be added in a second pass if needed.",
        ),
    ]
    for row_idx, (key, value) in enumerate(readme_rows, start=2):
        readme.write(row_idx, 0, key, key_fmt)
        readme.write(row_idx, 1, value)

    write_sheet(
        workbook,
        "Source Inventory",
        ["source_file", "source_url", "sheet", "status", "rows_scanned", "records_imported", "notes"],
        sheet_stats,
        {"source_file": 34, "source_url": 58, "sheet": 22, "status": 20, "notes": 90},
    )
    write_sheet(
        workbook,
        "Canonical Businesses",
        CANONICAL_COLUMNS,
        canonical_rows,
        {
            "canonical_id": 15,
            "business_name": 34,
            "normalized_name": 28,
            "category": 28,
            "source_family": 24,
            "address": 38,
            "source_files": 42,
            "source_refs": 70,
            "automatic_merge_basis": 60,
            "notes": 45,
        },
    )
    write_sheet(
        workbook,
        "Duplicate Groups",
        [
            "merge_group_id",
            "group_size",
            "primary_business_name",
            "city",
            "state",
            "automatic_merge_basis",
            "record_ids",
            "source_files",
            "phones",
            "emails",
            "licenses",
            "names_seen",
        ],
        duplicate_groups,
        {
            "primary_business_name": 34,
            "automatic_merge_basis": 60,
            "record_ids": 55,
            "source_files": 42,
            "names_seen": 45,
        },
    )
    write_sheet(
        workbook,
        "Duplicate Members",
        [
            "merge_group_id",
            "record_id",
            "business_name",
            "source_file",
            "source_tab",
            "source_row",
            "license_number",
            "phone",
            "email",
            "address",
            "city",
            "state",
        ],
        duplicate_members,
        {"business_name": 34, "source_file": 34, "address": 42, "email": 28},
    )
    write_sheet(
        workbook,
        "Needs Review",
        [
            "review_group_id",
            "reason",
            "canonical_id",
            "record_id",
            "business_name",
            "city",
            "state",
            "phone",
            "email",
            "address",
            "source_file",
            "source_tab",
            "source_row",
        ],
        review_rows,
        {"reason": 58, "business_name": 34, "address": 42, "source_file": 34},
    )
    write_sheet(
        workbook,
        "Excluded Screenshot Sheets",
        ["source_file", "source_url", "reason"],
        EXCLUDED_SCREENSHOT_SHEETS,
        {"source_file": 54, "source_url": 70, "reason": 58},
    )
    workbook.close()


def main():
    records, sheet_stats = build_records()
    canonical_rows, duplicate_groups, duplicate_members, review_rows = dedupe_records(records)
    write_workbook(canonical_rows, duplicate_groups, duplicate_members, review_rows, sheet_stats, len(records))

    imported_files = {
        row["source_file"]
        for row in sheet_stats
        if row.get("status") in {"imported", "no_business_rows"} and row.get("records_imported", 0)
    }
    stats = {
        "output_path": str(OUTPUT),
        "raw_business_rows_imported": len(records),
        "canonical_business_rows": len(canonical_rows),
        "automatic_duplicate_groups": len(duplicate_groups),
        "duplicate_member_rows": len(duplicate_members),
        "needs_review_rows": len(review_rows),
        "source_files_imported": len(imported_files),
        "source_sheets_inventory_rows": len(sheet_stats),
        "excluded_screenshot_sheets": len(EXCLUDED_SCREENSHOT_SHEETS),
        "identified_not_imported": len(IDENTIFIED_NOT_IMPORTED),
    }
    STATS_OUTPUT.write_text(json.dumps(stats, indent=2), encoding="utf-8")
    print(json.dumps(stats, indent=2))


if __name__ == "__main__":
    main()
