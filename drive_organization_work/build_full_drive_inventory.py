from __future__ import annotations

import os
import re
import sqlite3
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill


DRIVEFS_ROOT = Path(r"C:\Users\flavo\AppData\Local\Google\DriveFS")
ACCOUNT_IDS = ["106226209785867920657", "114146440822118596242"]
OUT_DIR = Path("drive_organization_work")
OUT_PATH = OUT_DIR / "Google Drive Full Inventory - 2026-07-05.xlsx"

FOLDER_MIME = "application/vnd.google-apps.folder"


def connect_readonly(path: Path) -> sqlite3.Connection:
    uri = "file:" + str(path).replace("\\", "/") + "?mode=ro"
    return sqlite3.connect(uri, uri=True)


def email_from_properties(con: sqlite3.Connection, fallback: str) -> str:
    rows = con.execute("select property, value from properties").fetchall()
    props = {key: value for key, value in rows}
    blob = props.get("driveway_account") or b""
    if isinstance(blob, str):
        blob = blob.encode("utf-8", "ignore")
    matches = re.findall(rb"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}Z?", blob)
    if matches:
        return matches[0].decode("utf-8", "replace").rstrip("Z")
    return fallback


def ms_to_utc(value: int | None) -> str:
    if not value:
        return ""
    try:
        return datetime.fromtimestamp(value / 1000, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    except (OverflowError, OSError, ValueError):
        return ""


def url_for(item_id: str, mime_type: str, is_folder: bool) -> str:
    if not item_id:
        return ""
    if is_folder or mime_type == FOLDER_MIME:
        return f"https://drive.google.com/drive/folders/{item_id}"
    if mime_type == "application/vnd.google-apps.document":
        return f"https://docs.google.com/document/d/{item_id}"
    if mime_type == "application/vnd.google-apps.spreadsheet":
        return f"https://docs.google.com/spreadsheets/d/{item_id}"
    if mime_type == "application/vnd.google-apps.presentation":
        return f"https://docs.google.com/presentation/d/{item_id}"
    if mime_type == "application/vnd.google-apps.form":
        return f"https://docs.google.com/forms/d/{item_id}"
    return f"https://drive.google.com/file/d/{item_id}"


def extension_for(name: str, is_folder: bool) -> str:
    if is_folder:
        return ""
    suffix = Path(name).suffix.lower()
    return suffix[1:] if suffix.startswith(".") else suffix


def protected_status(path: str) -> str:
    lowered = "/" + path.replace("\\", "/").lower().strip("/") + "/"
    reasons: list[str] = []
    if lowered.startswith("/my computer/"):
        reasons.append("synced computer tree")
    if "/aaatradercorner/" in lowered:
        reasons.append("AAATraderCorner repo/workspace tree")
    if "/repositories- do not share/" in lowered:
        reasons.append("Repositories- DO NOT SHARE")
    if "/.git/" in lowered:
        reasons.append("git metadata")
    for marker in ["/node_modules/", "/.venv/", "/venv/", "/__pycache__/", "/dist/", "/build/"]:
        if marker in lowered:
            reasons.append(marker.strip("/"))
            break
    if not reasons:
        return ""
    return "Protected: " + "; ".join(dict.fromkeys(reasons))


def header_row(ws, values: list[str]):
    fill = PatternFill("solid", fgColor="E8EAED")
    font = Font(bold=True, color="000000")
    cells = []
    for value in values:
        cell = WriteOnlyCell(ws, value=value)
        cell.fill = fill
        cell.font = font
        cells.append(cell)
    return cells


def safe_cell(value):
    if isinstance(value, str) and len(value) > 32000:
        return value[:31950] + " [truncated]"
    return value


def load_account(account_id: str) -> dict:
    db_path = DRIVEFS_ROOT / account_id / "metadata_sqlite_db"
    con = connect_readonly(db_path)
    email = email_from_properties(con, account_id)

    item_rows = con.execute(
        """
        select stable_id, id, local_title, trashed, is_tombstone, is_owner,
               mime_type, is_folder, modified_date, shared_with_me_date,
               viewed_by_me_date, file_size
        from items
        """
    ).fetchall()

    parent_rows = con.execute(
        "select item_stable_id, parent_stable_id from stable_parents"
    ).fetchall()
    con.close()

    items = {
        row[0]: {
            "stable_id": row[0],
            "id": row[1],
            "name": row[2] or "",
            "trashed": bool(row[3]),
            "is_tombstone": bool(row[4]),
            "is_owner": bool(row[5]),
            "mime_type": row[6] or "",
            "is_folder": bool(row[7]),
            "modified": row[8],
            "shared": row[9],
            "viewed": row[10],
            "size": row[11] or 0,
        }
        for row in item_rows
    }
    parent_map = {child: parent for child, parent in parent_rows}

    path_cache: dict[int, str] = {}

    def build_path(stable_id: int) -> str:
        if stable_id in path_cache:
            return path_cache[stable_id]
        seen: set[int] = set()
        chain: list[int] = []
        current = stable_id
        while current in items and current not in path_cache and current not in seen:
            seen.add(current)
            chain.append(current)
            parent = parent_map.get(current)
            if parent is None:
                break
            current = parent
        prefix = path_cache.get(current, "")
        parts = [items[sid]["name"] for sid in reversed(chain) if sid in items]
        if prefix and (not parts or prefix != parts[0]):
            full_parts = [prefix] + list(reversed(parts))
        else:
            full_parts = list(reversed(parts))
        path = ""
        for sid in reversed(chain):
            name = items[sid]["name"]
            if not path:
                parent = parent_map.get(sid)
                if parent in path_cache:
                    parent_path = path_cache[parent]
                    path = f"{parent_path}/{name}" if parent_path else name
                else:
                    path = name
            else:
                path = f"{path}/{name}"
            path_cache[sid] = path
        if stable_id not in path_cache:
            path_cache[stable_id] = "/".join(full_parts)
        return path_cache[stable_id]

    return {
        "account_id": account_id,
        "email": email,
        "items": items,
        "parent_map": parent_map,
        "build_path": build_path,
    }


def build_inventory() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook(write_only=True)

    ws = wb.create_sheet("Inventory")
    inventory_headers = [
        "account_email",
        "account_profile_id",
        "item_type",
        "name",
        "path",
        "top_folder",
        "parent_name",
        "drive_id",
        "parent_drive_id",
        "mime_type",
        "size_bytes",
        "modified_utc",
        "viewed_by_me_utc",
        "shared_with_me_utc",
        "owner_flag",
        "trashed",
        "url",
        "extension",
        "depth",
        "protected_status",
    ]
    ws.append(header_row(ws, inventory_headers))

    summary = {
        "account_records": Counter(),
        "account_unique_ids": defaultdict(set),
        "record_count": 0,
        "folder_count": 0,
        "file_count": 0,
        "trashed_count": 0,
        "tombstone_count": 0,
        "protected_count": 0,
        "mime_counts": Counter(),
        "top_folder_counts": Counter(),
        "duplicate_drive_ids": Counter(),
    }
    drive_id_accounts: dict[str, set[str]] = defaultdict(set)
    drive_id_rows: dict[str, list[tuple[str, str, str, str]]] = defaultdict(list)

    accounts = [load_account(account_id) for account_id in ACCOUNT_IDS]

    for account in accounts:
        items = account["items"]
        parent_map = account["parent_map"]
        build_path = account["build_path"]
        email = account["email"]
        account_id = account["account_id"]

        for stable_id in sorted(items):
            item = items[stable_id]
            path = build_path(stable_id)
            parts = [part for part in path.split("/") if part]
            top_folder = parts[1] if len(parts) > 1 else parts[0] if parts else ""
            parent_stable = parent_map.get(stable_id)
            parent = items.get(parent_stable) if parent_stable else None
            parent_name = parent["name"] if parent else ""
            parent_drive_id = parent["id"] if parent else ""
            item_type = "folder" if item["is_folder"] else "file"
            status = protected_status(path)
            url = url_for(item["id"], item["mime_type"], item["is_folder"])

            row = [
                email,
                account_id,
                item_type,
                item["name"],
                path,
                top_folder,
                parent_name,
                item["id"],
                parent_drive_id,
                item["mime_type"],
                item["size"],
                ms_to_utc(item["modified"]),
                ms_to_utc(item["viewed"]),
                ms_to_utc(item["shared"]),
                "yes" if item["is_owner"] else "no",
                "yes" if item["trashed"] else "no",
                url,
                extension_for(item["name"], item["is_folder"]),
                max(len(parts) - 1, 0),
                status,
            ]
            ws.append([safe_cell(value) for value in row])

            summary["record_count"] += 1
            summary["account_records"][email] += 1
            summary["account_unique_ids"][email].add(item["id"])
            summary["folder_count"] += 1 if item["is_folder"] else 0
            summary["file_count"] += 0 if item["is_folder"] else 1
            summary["trashed_count"] += 1 if item["trashed"] else 0
            summary["tombstone_count"] += 1 if item["is_tombstone"] else 0
            summary["protected_count"] += 1 if status else 0
            summary["mime_counts"][item["mime_type"]] += 1
            summary["top_folder_counts"][(email, top_folder)] += 1
            drive_id_accounts[item["id"]].add(email)
            if len(drive_id_rows[item["id"]]) < 5:
                drive_id_rows[item["id"]].append((email, item["name"], path, item["mime_type"]))

    unique_drive_ids = len(drive_id_accounts)
    duplicate_ids = {k: v for k, v in drive_id_accounts.items() if len(v) > 1}

    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(header_row(ws_summary, ["metric", "value"]))
    summary_rows = [
        ("scan_source", "Google Drive for Desktop DriveFS metadata SQLite cache"),
        ("generated_at", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")),
        ("account_cache_records", summary["record_count"]),
        ("unique_drive_ids", unique_drive_ids),
        ("duplicate_drive_ids_seen_in_multiple_accounts", len(duplicate_ids)),
        ("files", summary["file_count"]),
        ("folders", summary["folder_count"]),
        ("trashed_records_included", summary["trashed_count"]),
        ("tombstone_records_included", summary["tombstone_count"]),
        ("protected_or_repo_related_records", summary["protected_count"]),
        ("coverage_note", "Includes DriveFS metadata from both local account profiles; no files were moved, renamed, deleted, or downloaded."),
        ("sync_note", "Rows under My Computer or repo-like paths are marked protected so synced repositories can stay as-is."),
    ]
    for key, value in summary_rows:
        ws_summary.append([key, value])
    ws_summary.append([])
    ws_summary.append(header_row(ws_summary, ["account_email", "account_profile_id", "records", "unique_drive_ids"]))
    for account in accounts:
        email = account["email"]
        ws_summary.append([
            email,
            account["account_id"],
            summary["account_records"][email],
            len(summary["account_unique_ids"][email]),
        ])

    ws_mime = wb.create_sheet("Mime Summary")
    ws_mime.append(header_row(ws_mime, ["mime_type", "record_count"]))
    for mime_type, count in summary["mime_counts"].most_common():
        ws_mime.append([mime_type, count])

    ws_top = wb.create_sheet("Top Folders")
    ws_top.append(header_row(ws_top, ["account_email", "top_folder", "record_count"]))
    for (email, top_folder), count in summary["top_folder_counts"].most_common(1000):
        ws_top.append([email, top_folder, count])

    ws_dupes = wb.create_sheet("Duplicate Drive IDs")
    ws_dupes.append(header_row(ws_dupes, ["drive_id", "accounts", "sample_name", "sample_path", "mime_type"]))
    for drive_id in sorted(duplicate_ids):
        for email, name, path, mime_type in drive_id_rows[drive_id]:
            ws_dupes.append([drive_id, ", ".join(sorted(duplicate_ids[drive_id])), name, path, mime_type])

    ws_rules = wb.create_sheet("Rules")
    ws_rules.append(header_row(ws_rules, ["rule", "notes"]))
    rules = [
        ("No destructive actions", "This workbook is an inventory only. It does not move, rename, delete, or merge Drive files."),
        ("Repo/synced folders", "Anything under My Computer, AAATraderCorner, Repositories- DO NOT SHARE, .git, node_modules, venv, .venv, __pycache__, dist, or build is marked protected."),
        ("Every visible DriveFS item", "The Inventory tab includes both file and folder records from the local DriveFS metadata cache, including trashed/tombstone flags where present."),
        ("Unique ID accounting", "Summary distinguishes account-cache records from unique Drive IDs because the same Drive ID can appear in more than one account cache."),
        ("Next organization pass", "Use protected_status, top_folder, mime_type, and path filters before deciding what to archive, merge, or leave untouched."),
    ]
    for rule, note in rules:
        ws_rules.append([rule, note])

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH.resolve()}")
    print(f"records={summary['record_count']} unique_drive_ids={unique_drive_ids} duplicates={len(duplicate_ids)}")


if __name__ == "__main__":
    build_inventory()
